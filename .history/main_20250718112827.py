from flask import Flask, request, jsonify, render_template
import cohere
from dotenv import load_dotenv
import os
import pandas as pd
from openpyxl import load_workbook

# Load environment variables
load_dotenv()
api_key = os.getenv("COHERE_API_KEY")
co = cohere.Client(api_key)

# Load system prompt
with open("prompt_template.txt", "r") as f:
    system_prompt = f.read()

# Initialize Flask App
app = Flask(__name__)

EXCEL_FILE = "submissions.xlsx"  # Define file name

# Append data to Excel file
def append_to_excel(data_dict):
    df = pd.DataFrame([data_dict])

    if not os.path.exists(EXCEL_FILE):
        # Create new Excel file if it doesn't exist
        df.to_excel(EXCEL_FILE, index=False)
    else:
        # Append to existing Excel without header
        with pd.ExcelWriter(EXCEL_FILE, engine='openpyxl', mode='a', if_sheet_exists='overlay') as writer:
            workbook = load_workbook(EXCEL_FILE)
            sheet = workbook.active
            start_row = sheet.max_row
            df.to_excel(writer, index=False, header=False, startrow=start_row)

@app.route("/", methods=["GET"])
def index():
    return render_template("index.html")

@app.route("/parse-form", methods=["POST"])
def parse_form():
    try:
        data = request.get_json()
        user_prompt = data.get("prompt", "").strip()
        full_prompt = f"{system_prompt}\n\nUser Input:\n{user_prompt}"

        response = co.generate(
            model='command-r-plus',
            prompt=full_prompt,
            max_tokens=300,
            temperature=0.2
        )

        result_text = response.generations[0].text.strip()

        # Convert result text into dictionary
        result_lines = [line for line in result_text.split("\n") if ":" in line]
        result_dict = {}
        for line in result_lines:
            key, value = line.split(":", 1)
            result_dict[key.strip()] = value.strip()

        append_to_excel(result_dict)

        return jsonify({
            "status": "success",
            "result": result_text
        })

    except Exception as e:
        return jsonify({
            "status": "error",
            "message": str(e)
        }), 500

if __name__ == "__main__":
    app.run(debug=True)
